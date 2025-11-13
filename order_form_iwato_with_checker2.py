import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins

# === ファイル設定 ===
BASE = r"C:\Users\mokal\OneDrive\デスクトップ\My python"
INPUT = os.path.join(BASE, "検収記録簿 (1).xlsx")

# === Excel読み込み ===
df = pd.read_excel(INPUT, header=0)

# === 列名整理 ===
df = df.rename(columns={
    "使用日": "使用日",
    "朝昼夕": "朝昼夕",
    "仕入先": "仕入先",
    "食品名": "食品名",
    "単位": "単位",
    "入所者": "入所者",
    "職員": "職員",
    "ユーハウス入所者": "ユーハウス入所者",
    "備考欄": "備考欄"
})

# === 欠損補完 ===
for c in ["使用日", "仕入先", "食品名"]:
    df[c] = df[c].ffill()

# 使用日は並び替えに備えて一応日付型に（文字のままでも可）
df["使用日"] = pd.to_datetime(df["使用日"], errors="ignore")

# === 数値変換 ===
for c in ["入所者", "職員"]:
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

# === 出力列順設定 ===
keep_cols = [
    "使用日", "食品名", "入所者", "単位", "職員",
    "鮮度", "品温", "異物", "包装", "期限", "備考欄", "納品時間", "検収者"
]

# === 空列追加 ===
for c in keep_cols:
    if c not in df.columns:
        df[c] = ""

# === 出力対象（仕入先ごと） ===
suppliers = df["仕入先"].dropna().unique()

def apply_style(ws):
    """全体のフォントや罫線・列幅など"""
    font_body = Font(name="ＭＳ ゴシック", size=22)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    header_row = 6  # DataFrameのヘッダー行

    # --- 6行目のヘッダー：サイズ18、中央寄せ ---
    for cell in ws[header_row]:
        cell.font = Font(name="ＭＳ ゴシック", size=18, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # --- データ行（7行目以降） ---
    for row in ws.iter_rows(min_row=header_row + 1):
        for c in row:
            c.font = font_body
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border

    # --- 行の高さ ---
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 30

    # --- 列幅設定 ---
    ws.column_dimensions["A"].width = 15.18
    ws.column_dimensions["B"].width = 60.09   # 食品名
    for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws.column_dimensions[col].width = 15.18

    # --- B列に「縮小してセルに合わせる」を設定 ---
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            old_align = cell.alignment
            cell.alignment = Alignment(
                horizontal=old_align.horizontal,
                vertical=old_align.vertical,
                wrap_text=old_align.wrap_text,
                shrink_to_fit=True
            )

    # --- 用紙設定 ---
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

def create_header(ws, supplier):
    """FAX注文書のヘッダー"""

    # A3:B3 は結合したまま「○○御中」
    ws.merge_cells("A3:B3")
    ws["A3"] = f"{supplier}　御中"
    ws["A3"].font = Font(name="ＭＳ ゴシック", size=28, bold=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="bottom")

    # タイトルは B1セルのみ（セル結合なし）
    ws["B1"] = "注文書（介護老人福祉施設いわと）"
    ws["B1"].font = Font(name="ＭＳ ゴシック", size=26, bold=True)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

    # (有) ハートミールを K2 に配置（結合しない）
    ws["K2"] = "(有) ハートミール"
    ws["K2"].font = Font(name="ＭＳ ゴシック", size=24, bold=True)
    ws["K2"].alignment = Alignment(horizontal="right", vertical="center")

# === 出力処理 ===
for supplier in suppliers:
    sub = df[df["仕入先"] == supplier].copy()

    # 集計（使用日＋食品名＋単位でグループ化）
    sub = sub.groupby(
        ["使用日", "食品名", "単位"],
        as_index=False
    )[["入所者", "職員"]].sum()

    # 空列補完
    for c in keep_cols:
        if c not in sub.columns:
            sub[c] = ""

    # 並び替え：使用日 → 食品名
    sub = sub.sort_values(["使用日", "食品名"])
    sub = sub[keep_cols]

    # ★ 使用日の重複を空欄にする（A列＝使用日だけ）
    sub["使用日"] = sub["使用日"].mask(sub["使用日"].duplicated(), "")

    # 出力ファイル作成
    out_path = os.path.join(BASE, f"注文書_{supplier}_いわと.xlsx")
    sub.to_excel(out_path, index=False, startrow=5)

    wb = load_workbook(out_path)
    ws = wb.active

    apply_style(ws)
    create_header(ws, supplier)

    wb.save(out_path)

    print(f"✅ {supplier} の注文書を作成 → {out_path}")

print("\n🎉 介護老人福祉施設いわと 用 注文書の出力が完了しました！（検収者列を追加）")